import os
import argparse
import pandas as pd
from sqlalchemy import create_engine, text
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
import numpy as np
# utils
def ensure_dirs():
    os.makedirs("dv-assignment/charts", exist_ok=True)
    os.makedirs("dv-assignment/exports", exist_ok=True)

def mk_engine(a):
    dsn = f"postgresql+psycopg2://{a.user}:{a.password}@{a.host}:{a.port}/{a.dbname}"
    return create_engine(dsn, future=True)

def run_sql(engine, path, label):
    with engine.connect() as con:
        sql = open(path, "r", encoding="utf-8").read()
        df = pd.read_sql(text(sql), con)
    print(f"[OK] {label}: {len(df):,} rows")
    return df

def save_png(fig, name, title):
    path = f"dv-assignment/charts/{name}.png"
    fig.savefig(path, bbox_inches="tight")
    print(f"[CHART] {name}.png — {title}")

# charts
def pie_investor_types(df):
    title = "Распределение типов инвесторов по участию в сделках"
    fig, ax = plt.subplots()
    s = df.set_index("investor_type")["deals"]
    s.plot.pie(ax=ax, autopct="%1.1f%%", ylabel="")
    ax.set_title(title)
    save_png(fig, "pie_investor_types", title)

def bar_top_buyers(df):
    title = "ТОП-10 покупателей M&A по числу сделок"
    fig, ax = plt.subplots()
    ax.bar(df["buyer"], df["deals"])
    ax.set_title(title); ax.set_xlabel("Покупатель"); ax.set_ylabel("Сделки")
    plt.xticks(rotation=45, ha="right")
    save_png(fig, "bar_top_buyers", title)

def barh_countries_raised(df):
    title = "ТОП-10 стран по сумме привлечений (2005–2015, с инвесторами)"
    fig, ax = plt.subplots()
    ax.barh(df["country"], df["raised_usd"]); ax.invert_yaxis()
    ax.set_title(title); ax.set_xlabel("Сумма, USD"); ax.set_ylabel("Страна")
    save_png(fig, "barh_countries_raised", title)

def line_top5_investors(df):
    title = "Динамика Total Raised по годам — ТОП-5 инвесторов"
    fig, ax = plt.subplots()
    for inv, g in df.groupby("investor"):
        ax.plot(g["year"], g["raised_usd"], label=inv)
    ax.set_title(title); ax.set_xlabel("Год"); ax.set_ylabel("Сумма, USD")
    ax.legend(fontsize=8)
    save_png(fig, "line_top5_investors", title)

def hist_seriesa_usa(df):
    title = "Распределение размера Series A в США (только раунды с инвесторами)"
    fig, ax = plt.subplots()
    ax.hist(df["raised_amount_usd"], bins=30)
    ax.set_title(title); ax.set_xlabel("Raised, USD"); ax.set_ylabel("Количество раундов")
    save_png(fig, "hist_seriesa_usa", title)

def scatter_funding_vs_acq(df):
    title = "Total funding vs. число поглощений как цель (по компаниям)"
    fig, ax = plt.subplots()
    ax.scatter(df["total_raised"], df["acquisitions_as_target"], alpha=0.6)
    ax.set_title(title); ax.set_xlabel("Total raised, USD"); ax.set_ylabel("Acquisitions as target, count")
    save_png(fig, "scatter_funding_vs_acq", title)



def plotly_country_year(df_anim):
    df = df_anim.copy()
    df["deals"] = df["deals"].clip(lower=1)
    df["avg_raised"] = df["avg_raised"].clip(lower=1)

    x_lo, x_hi = np.quantile(df["deals"], [0.05, 0.98])
    y_lo, y_hi = np.quantile(df["avg_raised"], [0.05, 0.98])

    p95 = np.quantile(df["raised_usd"], 0.95)
    df["size_val"] = np.clip(df["raised_usd"], None, p95)

    size_max = 120          # было 60 → делаем крупнее
    sizeref = 2.0 * p95 / (size_max ** 2)

    fig = px.scatter(
        df, x="deals", y="avg_raised",
        animation_frame="year", animation_group="country",
        color="country",
        size="size_val", size_max=size_max,
        hover_name="country",
        title="Страны: сделки vs средний размер раунда по годам (топ-10)"
    )

    fig.update_xaxes(type="log", range=[np.log10(x_lo), np.log10(x_hi)], title="deals (log)")
    fig.update_yaxes(type="log", range=[np.log10(y_lo), np.log10(y_hi)], title="avg_raised, USD (log)")

    fig.update_traces(
        marker=dict(
            sizemode="area",
            sizeref=sizeref,
            sizemin=6,        # минимальный видимый радиус
            line=dict(width=0.5),
            opacity=0.9
        ),
        hovertemplate="<b>%{hovertext}</b><br>"
                      "deals: %{x:.0f}<br>"
                      "avg raised: %{y:$,.0f}<br>"
                      "total raised (p95-cap): %{marker.size:$,.0f}<extra></extra>"
    )

    fig.update_layout(template="simple_white", transition={'duration': 300},
                      margin=dict(l=40, r=20, t=60, b=40))
    fig.show()



# ==== Экспорт в Excel с форматированием ====
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule

def export_to_excel(dataframes: dict[str, pd.DataFrame], filename: str):

    out_dir = Path("dv-assignment/exports")
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / filename

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in dataframes.items():
            sheet_name = (sheet or "Sheet")[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path)
    total_rows = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        ws.freeze_panes = "B2"

        ws.auto_filter.ref = ws.dimensions

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col[: min(len(col), 200)]:
                v = cell.value
                v = "" if v is None else str(v)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = min(max(10, int(max_len * 0.9)), 60)

        header = [c.value for c in ws[1]]
        rows_count = ws.max_row - 1 if ws.max_row > 1 else 0
        total_rows += rows_count

        numeric_cols = []
        for j in range(1, len(header) + 1):
            values = []
            for i in range(2, min(ws.max_row, 101)):
                v = ws.cell(row=i, column=j).value
                if isinstance(v, (int, float)):
                    values.append(v)
                elif isinstance(v, str):
                    try:
                        float(v.replace(",", ""))
                        values.append(0.0)
                    except Exception:
                        pass
            if values:
                numeric_cols.append(j)

        # градиент и min/max для каждой числовой колонки
        for col_idx in numeric_cols:
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            data_range = f"{col_letter}2:{col_letter}{ws.max_row}"

            # 3-цветный градиент
            grad = ColorScaleRule(
                start_type="min", start_color="FFAA0000",   # красный
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00",  # жёлтый
                end_type="max", end_color="FF00AA00"        # зелёный
            )
            ws.conditional_formatting.add(data_range, grad)

            # MIN и MAX (тонкая подсветка поверх градиента)
            min_fill = PatternFill("solid", fgColor="FFB3D1FF")
            max_fill = PatternFill("solid", fgColor="FFFFC04D")

            ws.conditional_formatting.add(
                data_range,
                FormulaRule(formula=[f'{col_letter}2=MIN({data_range})'], fill=min_fill),
            )
            ws.conditional_formatting.add(
                data_range,
                FormulaRule(formula=[f'{col_letter}2=MAX({data_range})'], fill=max_fill),
            )

    wb.save(path)

    print(f"Создан файл {path.name}, {len(wb.sheetnames)} листа(ов), {total_rows} строк.")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--host", default="localhost")
    ap.add_argument("--port", type=int, default=5432)
    ap.add_argument("--dbname", default="dv_project")
    ap.add_argument("--user", default="postgres")
    ap.add_argument("--password", default="0000")
    args = ap.parse_args()

    ensure_dirs()
    engine = mk_engine(args)

    base = "dv-assignment/sql/assignment2/"

    df_pie  = run_sql(engine, base + "pie_investor_types.sql",        "PIE")
    df_bar  = run_sql(engine, base + "bar_top_buyers.sql",            "BAR")
    df_bh   = run_sql(engine, base + "barh_countries_raised.sql",     "BARH")
    df_line = run_sql(engine, base + "line_top5_investors_by_year.sql","LINE")
    df_hist = run_sql(engine, base + "hist_seriesa_usa.sql",          "HIST")
    df_scat = run_sql(engine, base + "scatter_funding_vs_acq.sql",    "SCATTER")
    df_anim = run_sql(engine, base + "plotly_country_year.sql",       "PLOTLY")

    pie_investor_types(df_pie)
    bar_top_buyers(df_bar)
    barh_countries_raised(df_bh)
    line_top5_investors(df_line)
    hist_seriesa_usa(df_hist)
    scatter_funding_vs_acq(df_scat)

    print("[PLOTLY] Откроется интерактивный график; закрой окно для продолжения.")
    plotly_country_year(df_anim)

    export_to_excel({
        "investor_types": df_pie,
        "top_buyers": df_bar,
        "countries_2005_2015": df_bh,
        "top5_inv_by_year": df_line,
        "seriesA_USA": df_hist,
        "funding_vs_acq": df_scat,
    }, "assignment2_report.xlsx")

if __name__ == "__main__":
    main()
